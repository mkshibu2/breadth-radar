#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════╗
║  NSE VCP SCREENER — Minervini / O'Neil / Stage-2 Method              ║
║  Version 2.4  |  Free data via Yahoo Finance  |  by Shibu            ║
║                                                                      ║
║  Scores every Nifty 500 stock on:                                    ║
║    • Minervini Trend Template (Stage 2 confirmation)                 ║
║    • Prior Advance (momentum qualification)                          ║
║    • VCP: Drawdown, Volume dry-up, ATR contraction, Tightness        ║
║    • True VCP pullback sequence detection                            ║
║    • Relative Strength Proxy (vs universe)                           ║
║    • Pocket Pivot detection                                          ║
║    • Breakout Readiness Score                                        ║
║    • Final Composite Score + Grade + Setup Type                      ║
║                                                                      ║
║  Install: pip install yfinance pandas openpyxl                       ║
║  Run:     python dhan_scanner.py                                     ║
║  Output:  scanner_results.html  - to NSE Breadth Radar               ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import yfinance as yf
import pandas as pd
import numpy as np
import time, logging, webbrowser, warnings, sys, signal
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings('ignore')
logging.basicConfig(level=logging.WARNING, format='%(message)s')

# ═══════════════════════════════════════════════════════════════════════
#  CTRL+C OVERRIDE & CFFI POPUP SUPPRESSION
# ═══════════════════════════════════════════════════════════════════════
def silent_unraisable_hook(unraisable):
    if issubclass(unraisable.exc_type, KeyboardInterrupt):
        return
    sys.__unraisablehook__(unraisable)
sys.unraisablehook = silent_unraisable_hook

class GracefulInterruptHandler:
    def __init__(self):
        self.interrupted = False
        signal.signal(signal.SIGINT, self.handler)

    def handler(self, signum, frame):
        if self.interrupted:
            print("\n\n🛑 Force quitting...")
            sys.exit(1)
        self.interrupted = True
        print("\n\n🛑 [Ctrl+C Detected] Please wait... finishing the current stock download to avoid breaking the network connection. The report will generate right after!")

# ═══════════════════════════════════════════════════════════════════════
#  CONFIGURATION — edit these to tune the screener
# ═══════════════════════════════════════════════════════════════════════
CFG = {
    # Custom Links
    'tv_chart_url': 'https://in.tradingview.com/chart/fZ5fwxUh/',

    # Liquidity filters
    'min_avg_volume_30d':    200_000,
    'min_avg_turnover_crore': 5,
    'min_market_cap_crore':   500,

    # Prior advance thresholds
    'min_30d_return':  15.0,   # %
    'min_60d_return':  25.0,   # %
    'pref_90d_return': 30.0,   # %

    # 52-week high proximity
    'max_dist_52w_high': -25.0,  # % — reject below this

    # Consolidation
    'max_drawdown_30d': -25.0,   # % — reject below this

    # ATR contraction
    'atr_contraction_threshold': 0.85,

    # Scoring weights
    'vcp_weight':        0.60,
    'leadership_weight': 0.40,

    # Output
    'top_n': 20,
    'delay_between_stocks': 0.25,
}

# ═══════════════════════════════════════════════════════════════════════
#  NIFTY 500 SYMBOLS
# ═══════════════════════════════════════════════════════════════════════
NIFTY500 = [
    "360ONE","3MINDIA","ABB","ABBOTINDIA","ABCAPITAL","ABFRL","ACC",
    "AARTIIND","ADANIENT","ADANIGREEN","ADANIPORTS","ADANIPOWER",
    "AFFLE","AJANTPHARM","ALKEM","ALKYLAMINE","ALLCARGO",
    "AMBUJACEM","ANANDRATHI","ANGELONE","ANURAS","APARINDS",
    "APOLLOHOSP","APOLLOTYRE","APTUS","ARVINDFASN","ASAHIINDIA",
    "ASHOKLEY","ASIANPAINT","ASTRAL","ATGL","ATUL","AUROPHARMA",
    "AVANTIFEED","AXISBANK","BAJAJ-AUTO","BAJAJFINSV","BAJAJHLDNG",
    "BAJFINANCE","BALKRISIND","BALRAMCHIN","BANDHANBNK","BANKBARODA",
    "BANKINDIA","BATAINDIA","BEL","BERGEPAINT","BHARATFORG",
    "BHARTIARTL","BHEL","BIOCON","BIRLACORPN","BLUEDART",
    "BOSCHLTD","BPCL","BRIGADE","BSE","BSOFT","CAMS","CANFINHOME",
    "CANBK","CAPLIPOINT","CARBORUNIV","CASTROLIND","CCL","CDSL",
    "CENTURYPLY","CESC","CHAMBLFERT","CHOLAFIN","CIPLA",
    "COALINDIA","COCHINSHIP","COFORGE","COLPAL","CONCOR",
    "COROMANDEL","CREDITACC","CROMPTON","CSBBANK","CUB",
    "CUMMINSIND","CYIENT","DATAPATTNS","DBREALTY","DCMSHRIRAM",
    "DEEPAKNTR","DELHIVERY","DEVYANI","DIVISLAB","DIXON","DLF",
    "DMART","DODLA","DRREDDY","EIDPARRY","ELGIEQUIP","EMAMILTD",
    "ENDURANCE","ENGINERSIN","ESCORTS","EXIDEIND","FACT",
    "FINCABLES","FLUOROCHEM","FORTIS","FSL","GAIL","GALAXYSURF",
    "GILLETTE","GLAXO","GLENMARK","GNFC","GODFRYPHLP","GODREJCP",
    "GODREJIND","GODREJPROP","GRANULES","GRAPHITE","GRASIM",
    "GREENPANEL","GRINDWELL","GUJGASLTD","HAL","HAPPSTMNDS",
    "HAVELLS","HCLTECH","HDFCAMC","HDFCBANK","HDFCLIFE","HFCL",
    "HIKAL","HINDCOPPER","HINDPETRO","HINDUNILVR","HONAUT","HUDCO",
    "IBREALEST","ICICIBANK","ICICIGI","ICICIPRULI","IDBI",
    "IDFCFIRSTB","IEX","IGL","IIFL","INDHOTEL","INDIAMART",
    "INDIANB","INDIGO","INDUSINDBK","INDUSTOWER","INFY","IOB",
    "IOC","IPCALAB","IRB","IRCTC","IRFC","ITC","ITI","J&KBANK",
    "JBCHEPHARM","JKCEMENT","JKLAKSHMI","JMFINANCIL","JSWENERGY",
    "JSWSTEEL","JUBLFOOD","JUBLINGREA","JUSTDIAL","JYOTHYLAB",
    "KAJARIACER","KANSAINER","KEC","KFINTECH","KNRCON","KOTAKBANK",
    "KPIL","KRBL","KSCL","LATENTVIEW","LAURUSLABS","LICI",
    "LICHSGFIN","LINDEINDIA","LT","LTTS","LUPIN","LUXIND",
    "M&M","M&MFIN","MARICO","MARUTI","MASTEK","MAXHEALTH","MCX",
    "MEDANTA","METROPOLIS","MFSL","MGL","MPHASIS","MRF","MRPL",
    "MUTHOOTFIN","NATCOPHARM","NAUKRI","NAVINFLUOR","NBCC","NCC",
    "NESTLEIND","NHPC","NIACL","NLCINDIA","NMDC","NOCIL","NTPC",
    "NUVOCO","OBEROIRLTY","OIL","ONGC","ORIENTELEC","PAGEIND",
    "PERSISTENT","PETRONET","PFC","PGHH","PHOENIXLTD","PIDILITIND",
    "PIIND","PNBHOUSING","POLICYBZR","POLYCAB","POLYMED",
    "POONAWALLA","POWERGRID","RADICO","RAILTEL","RAINBOW",
    "RAJESHEXPO","RAMCOCEM","RATNAMANI","RAYMOND","RECLTD",
    "REDINGTON","RELIANCE","RITES","RKFORGE","RPOWER","SAFARI",
    "SANOFI","SCI","SHREECEM","SHRIRAMFIN","SIEMENS","SJVN",
    "SKFINDIA","SOBHA","SOLARA","SONACOMS","SRF","STARHEALTH",
    "STCINDIA","SUNPHARMA","SUNTV","SUPREMEIND","SUZLON",
    "SYMPHONY","TANLA","TATACOMM","TATACONSUM","TATAELXSI",
    "TATAPOWER","TATASTEEL","TATATECH","TCI","TEAMLEASE","TIINDIA",
    "TIMKEN","TITAGARH","TITAN","TORNTPHARM","TORNTPOWER","TRENT",
    "TRIDENT","TRIVENI","UCOBANK","UJJIVANSFB","ULTRACEMCO",
    "UNIONBANK","UPL","UTIAMC","VAIBHAVGBL","VBL","VEDL",
    "VINATIORGA","VOLTAS","VSTIND","WABAG","WELCORP","WESTLIFE",
    "WHIRLPOOL","WIPRO","YESBANK","ZEEL","ZYDUSLIFE",
]

# ═══════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════
def safe(v, default=0.0):
    try:
        f = float(v)
        return default if (np.isnan(f) or np.isinf(f)) else f
    except: return default

def atr(df, period):
    """Average True Range over `period` days."""
    h, l, c = df['High'], df['Low'], df['Close']
    tr = pd.concat([h - l,
                    (h - c.shift()).abs(),
                    (l - c.shift()).abs()], axis=1).max(axis=1)
    return tr.rolling(period).mean().iloc[-1]

def ema(series, period):
    return series.ewm(span=period, adjust=False).mean()

def sma(series, period):
    return series.rolling(period).mean()

# ═══════════════════════════════════════════════════════════════════════
#  DATA FETCH
# ═══════════════════════════════════════════════════════════════════════
def fetch_data(symbol, days=400):
    """Fetch daily OHLCV + basic info from Yahoo Finance."""
    try:
        tk    = yf.Ticker(symbol + '.NS')
        end   = datetime.now()
        start = end - timedelta(days=days)
        df    = tk.history(start=start, end=end, interval='1d')
        if df is None or len(df) < 252:
            return None, {}
        df = df[['Open','High','Low','Close','Volume']].dropna()
        info = {}
        try:
            raw        = tk.fast_info
            info['market_cap'] = safe(getattr(raw, 'market_cap', 0)) / 1e7  # → crore
            info['company']    = getattr(raw, 'quote_type', symbol)
            raw2               = tk.info
            info['sector']     = raw2.get('sector', '—')
            info['industry']   = raw2.get('industry', '—')
            info['company']    = raw2.get('shortName', symbol)
        except: pass
        return df, info
    except: return None, {}

# ═══════════════════════════════════════════════════════════════════════
#  PART 2 — LIQUIDITY FILTER
# ═══════════════════════════════════════════════════════════════════════
def check_liquidity(df, info):
    avg_vol      = df['Volume'].iloc[-30:].mean()
    avg_turnover = (df['Close'].iloc[-30:] * df['Volume'].iloc[-30:]).mean() / 1e7  # crore
    market_cap   = safe(info.get('market_cap', 0))
    ok = (avg_vol      >= CFG['min_avg_volume_30d'] and
          avg_turnover >= CFG['min_avg_turnover_crore'] and
          (market_cap  >= CFG['min_market_cap_crore'] or market_cap == 0))
    return ok, round(avg_vol), round(avg_turnover, 2), round(market_cap)

# ═══════════════════════════════════════════════════════════════════════
#  PART 3 — MINERVINI TREND TEMPLATE
# ═══════════════════════════════════════════════════════════════════════
def trend_template(df):
    c      = df['Close']
    e10    = ema(c, 10).iloc[-1]
    e20    = ema(c, 20).iloc[-1]
    e50    = ema(c, 50).iloc[-1]
    s150   = sma(c, 150).iloc[-1]
    s200   = sma(c, 200).iloc[-1]
    close  = c.iloc[-1]

    conds = [
        close  > e10,    # Price above EMA10
        e10    > e20,    # EMA10 > EMA20
        e20    > e50,    # EMA20 > EMA50
        e50    > s150,   # EMA50 > SMA150
        s150   > s200,   # SMA150 > SMA200 (uptrend)
        close  > s150,   # Price > SMA150
        close  > s200,   # Price > SMA200
    ]
    score = sum(conds) / len(conds) * 35  # max 35 pts
    passed = sum(conds)
    return round(score, 1), passed, len(conds)

# ═══════════════════════════════════════════════════════════════════════
#  PART 4 — PRIOR ADVANCE
# ═══════════════════════════════════════════════════════════════════════
def prior_advance(df):
    c   = df['Close']
    cur = c.iloc[-1]
    def ret(n):
        idx = max(0, len(c) - n - 1)
        p   = c.iloc[idx]
        return round((cur / p - 1) * 100, 1) if p > 0 else 0.0

    r30 = ret(21)   # ~1 month trading days
    r60 = ret(42)   # ~2 months
    r90 = ret(63)   # ~3 months

    if   r90 >= 100: pts = 20
    elif r90 >= 70:  pts = 18
    elif r90 >= 50:  pts = 15
    elif r90 >= 30:  pts = 12
    elif r60 >= 30:  pts = 10
    elif r30 >= 15:  pts = 6
    else:            pts = 0

    qualifies = r30 >= CFG['min_30d_return'] or r60 >= CFG['min_60d_return']
    return r30, r60, r90, pts, qualifies

# ═══════════════════════════════════════════════════════════════════════
#  PART 5 — 52-WEEK HIGH PROXIMITY
# ═══════════════════════════════════════════════════════════════════════
def high_proximity(df):
    c       = df['Close']
    h252    = df['High'].iloc[-252:].max()
    l252    = df['Low'].iloc[-252:].min()
    cur     = c.iloc[-1]
    dist    = round((cur - h252) / h252 * 100, 1)

    if   dist >= -5:  pts = 20
    elif dist >= -10: pts = 15
    elif dist >= -15: pts = 10
    elif dist >= -20: pts = 5
    else:             pts = 0

    qualifies = dist >= CFG['max_dist_52w_high']
    return dist, pts, round(h252, 2), round(l252, 2), qualifies

# ═══════════════════════════════════════════════════════════════════════
#  PART 6 — CONSOLIDATION / DRAWDOWN
# ═══════════════════════════════════════════════════════════════════════
def consolidation(df):
    c        = df['Close']
    h30      = df['High'].iloc[-30:].max()
    cur      = c.iloc[-1]
    drawdown = round((cur - h30) / h30 * 100, 1)

    if   drawdown >= -5:   pts = 20
    elif drawdown >= -10:  pts = 16
    elif drawdown >= -15:  pts = 12
    elif drawdown >= -20:  pts = 8
    elif drawdown >= -25:  pts = 4
    else:                  pts = 0

    qualifies  = drawdown >= CFG['max_drawdown_30d']
    base_type  = ('Shelf'     if drawdown >= -5  else
                  'Flat Base' if drawdown >= -15 else
                  'Cup'       if drawdown >= -25 else 'Too Deep')
    return drawdown, pts, qualifies, base_type

# ═══════════════════════════════════════════════════════════════════════
#  PART 7 — VOLUME DRY-UP
# ═══════════════════════════════════════════════════════════════════════
def volume_dryup(df):
    v         = df['Volume']
    vol_r     = v.iloc[-10:].mean()
    vol_l     = v.iloc[-30:-10].mean()
    if vol_l <= 0: return 0.0, 0, False
    contraction = (vol_r / vol_l - 1) * 100
    dryup       = contraction < 0

    if   contraction <= -60: pts = 20
    elif contraction <= -40: pts = 16
    elif contraction <= -20: pts = 12
    elif contraction < 0:    pts = 6
    else:                    pts = 0

    return round(contraction, 1), pts, dryup

# ═══════════════════════════════════════════════════════════════════════
#  PART 8 — ATR CONTRACTION
# ═══════════════════════════════════════════════════════════════════════
def atr_contraction(df):
    if len(df) < 35: return 1.0, 0, False
    a10 = atr(df, 10)
    a30 = atr(df, 30)
    if a30 <= 0: return 1.0, 0, False
    ratio = round(a10 / a30, 3)

    if   ratio < 0.60: pts = 20
    elif ratio < 0.70: pts = 16
    elif ratio < 0.80: pts = 12
    elif ratio < 0.85: pts = 8
    else:              pts = 0

    contracting = ratio < CFG['atr_contraction_threshold']
    return ratio, pts, contracting

# ═══════════════════════════════════════════════════════════════════════
#  PART 9 — TIGHTNESS DETECTION
# ═══════════════════════════════════════════════════════════════════════
def tightness(df):
    h10 = df['High'].iloc[-10:].max()
    l10 = df['Low'].iloc[-10:].min()
    if l10 <= 0: return 99.0, 0
    tight = round((h10 - l10) / l10 * 100, 1)

    if   tight <= 5:  pts = 20
    elif tight <= 8:  pts = 16
    elif tight <= 10: pts = 12
    elif tight <= 15: pts = 6
    else:             pts = 0

    return tight, pts

# ═══════════════════════════════════════════════════════════════════════
#  PART 10 — TRUE VCP PULLBACK SEQUENCE
# ═══════════════════════════════════════════════════════════════════════
def vcp_pullbacks(df):
    c    = df['Close'].iloc[-90:].reset_index(drop=True)
    n    = len(c)
    troughs = []
    for i in range(5, n - 5):
        if c.iloc[i] == c.iloc[i-5:i+5].min():
            troughs.append((i, c.iloc[i]))

    peaks = []
    for i in range(5, n - 5):
        if c.iloc[i] == c.iloc[i-5:i+5].max():
            peaks.append((i, c.iloc[i]))

    pullbacks = []
    for t_idx, t_val in troughs:
        prior_peaks = [(p_i, p_v) for p_i, p_v in peaks if p_i < t_idx]
        if not prior_peaks: continue
        p_i, p_v = prior_peaks[-1]
        if p_v > 0:
            depth = (t_val - p_v) / p_v * 100
            pullbacks.append(round(abs(depth), 1))

    pullbacks = pullbacks[-4:]

    bonus = 0
    if len(pullbacks) >= 2:
        contractions = sum(1 for i in range(1, len(pullbacks))
                          if pullbacks[i] < pullbacks[i-1])
        ratio = contractions / (len(pullbacks) - 1)
        if   ratio == 1.0: bonus = 10
        elif ratio >= 0.67: bonus = 7
        elif ratio >= 0.5:  bonus = 4
        else:               bonus = 1

    return bonus, pullbacks

# ═══════════════════════════════════════════════════════════════════════
#  PART 11 — RS PROXY (computed after all stocks are scored)
# ═══════════════════════════════════════════════════════════════════════
def rs_proxy_raw(df):
    c   = df['Close']
    cur = c.iloc[-1]
    def ret(n):
        idx = max(0, len(c) - n - 1)
        p   = c.iloc[idx]
        return (cur / p - 1) * 100 if p > 0 else 0.0
    r63  = ret(63)
    r126 = ret(126)
    r252 = ret(252)
    composite = r63 * 0.4 + r126 * 0.3 + r252 * 0.3
    return composite, r63, r126, r252

def compute_rs_scores(results):
    composites = [r['_rs_composite'] for r in results]
    if not composites: return results
    mn, mx = min(composites), max(composites)
    rng    = mx - mn if mx != mn else 1
    for r in results:
        r['rs_score'] = round((r['_rs_composite'] - mn) / rng * 100, 1)
    return results

# ═══════════════════════════════════════════════════════════════════════
#  PART 12 — POCKET PIVOT DETECTION
# ═══════════════════════════════════════════════════════════════════════
def pocket_pivot(df):
    c      = df['Close']
    v      = df['Volume']
    e10_s  = ema(c, 10)

    if len(df) < 12: return False

    today_vol  = v.iloc[-1]
    today_cls  = c.iloc[-1]
    today_e10  = e10_s.iloc[-1]

    prior = df.iloc[-11:-1]
    down_days = prior[prior['Close'] < prior['Close'].shift(1).fillna(prior['Close'])]
    if down_days.empty: return False

    max_down_vol = down_days['Volume'].max()
    return bool(today_cls > today_e10 and today_vol > max_down_vol)

# ═══════════════════════════════════════════════════════════════════════
#  PART 13 — BREAKOUT READINESS SCORE
# ═══════════════════════════════════════════════════════════════════════
def breakout_readiness(df, dist_52w, tight_pct, atr_ratio, dryup, vol_contr):
    score  = 0
    c      = df['Close']
    v      = df['Volume']

    h30    = df['High'].iloc[-30:].max()
    cur    = c.iloc[-1]
    to_piv = (cur - h30) / h30 * 100
    if   to_piv >= -3:  score += 30
    elif to_piv >= -5:  score += 20
    elif to_piv >= -8:  score += 10

    v3_avg  = v.iloc[-3:].mean()
    v10_avg = v.iloc[-13:-3].mean()
    if v10_avg > 0 and v3_avg > v10_avg: score += 20

    if atr_ratio < 0.80: score += 20
    elif atr_ratio < 0.85: score += 10

    if   tight_pct <= 5: score += 20
    elif tight_pct <= 8: score += 14
    elif tight_pct <= 10: score += 7

    if dryup: score += 10

    return min(100, score)

# ═══════════════════════════════════════════════════════════════════════
#  PARTS 14-17 — SCORING & GRADING
# ═══════════════════════════════════════════════════════════════════════
def leadership_score(prox_pts, trend_pts, rs_score, avg_turnover):
    liq = min(15, avg_turnover / 10 * 15) if avg_turnover < 10 else 15
    rs_pts = rs_score * 0.30
    total = min(100, prox_pts + trend_pts + rs_pts + liq)
    return round(total, 1)

def vcp_score(adv_pts, dd_pts, vol_pts, atr_pts, tight_pts):
    return min(100, adv_pts + dd_pts + vol_pts + atr_pts + tight_pts)

def composite_score(vcp, leadership):
    return round(vcp * CFG['vcp_weight'] + leadership * CFG['leadership_weight'], 1)

def grade(score):
    if   score >= 90: return 'A+'
    elif score >= 80: return 'A'
    elif score >= 70: return 'B'
    elif score >= 60: return 'C'
    else:             return 'REJECT'

# ═══════════════════════════════════════════════════════════════════════
#  PART 18 — SETUP TYPE CLASSIFICATION
# ═══════════════════════════════════════════════════════════════════════
def setup_type(drawdown, tight_pct, atr_ratio, dist_52w, dryup, bos):
    if bos >= 60 and dist_52w >= -3 and tight_pct <= 8:
        return '🚀 Breakout Ready'
    elif drawdown >= -10 and tight_pct <= 8 and dryup:
        return '🎯 Mature VCP'
    elif drawdown >= -5:
        return '🌱 Early VCP'
    elif atr_ratio < 0.80:
        return '📐 Contraction'
    else:
        return '👀 Watch'

# ═══════════════════════════════════════════════════════════════════════
#  MAIN ANALYSIS FUNCTION
# ═══════════════════════════════════════════════════════════════════════
def analyse(symbol, df, info):
    c = df['Close']
    liq_ok, avg_vol, avg_turn, mktcap = check_liquidity(df, info)
    trend_pts, trend_passed, trend_total = trend_template(df)
    r30, r60, r90, adv_pts, adv_ok = prior_advance(df)
    if not adv_ok: return None, '< min advance'
    dist52, prox_pts, h52, l52, prox_ok = high_proximity(df)
    if not prox_ok: return None, 'too far from 52W high'
    drawdown, dd_pts, dd_ok, btype = consolidation(df)
    if not dd_ok: return None, 'drawdown too deep'

    vol_contr, vol_pts, dryup = volume_dryup(df)
    atr_ratio, atr_pts, atr_ok = atr_contraction(df)
    tight_pct, tight_pts = tightness(df)
    vcp_bonus, pullbacks = vcp_pullbacks(df)
    pp = pocket_pivot(df)
    rs_comp, rs63, rs126, rs252 = rs_proxy_raw(df)

    vcp_s  = vcp_score(adv_pts, dd_pts, vol_pts, atr_pts, tight_pts) + vcp_bonus
    vcp_s  = min(100, vcp_s)
    lead_s = 0

    bos = breakout_readiness(df, dist52, tight_pct, atr_ratio, dryup, vol_contr)
    stype = setup_type(drawdown, tight_pct, atr_ratio, dist52, dryup, bos)

    return {
        'symbol':       symbol,
        'company':      info.get('company', symbol),
        'sector':       info.get('sector', '—'),
        'industry':     info.get('industry', '—'),
        'market_cap':   mktcap,
        'current':      round(c.iloc[-1], 2),
        'h52':          h52,
        'l52':          l52,
        'avg_vol':      avg_vol,
        'avg_turn':     avg_turn,
        'r30':          r30,
        'r60':          r60,
        'r90':          r90,
        'dist52':       dist52,
        'drawdown':     drawdown,
        'base_type':    btype,
        'vol_contr':    vol_contr,
        'atr_ratio':    atr_ratio,
        'tight_pct':    tight_pct,
        'trend_passed': f'{trend_passed}/{trend_total}',
        'pullbacks':    ' → '.join(str(p)+'%' for p in pullbacks) or '—',
        'pocket_pivot': '✅ YES' if pp else '—',
        'vcp_score':    vcp_s,
        'lead_score':   0,
        'rs_score':     0,
        'final_score':  0,
        'grade':        '—',
        'setup':        stype,
        'bos':          bos,
        'adv_pts':      adv_pts,
        'prox_pts':     prox_pts,
        'trend_pts':    trend_pts,
        'avg_turn_liq': avg_turn,
        '_rs_composite': rs_comp,
    }, None

# ═══════════════════════════════════════════════════════════════════════
#  HTML & JS GENERATION
# ═══════════════════════════════════════════════════════════════════════

UI_CSS = r"""
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{background:#060d18;color:#e2e8f5;font-family:system-ui,-apple-system,sans-serif;min-height:100vh}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-thumb{background:#2d3f5a;border-radius:3px}
::-webkit-scrollbar-track{background:#0d1929}
.tab{display:none}.tab.active{display:block}
.tab-btn{background:none;border:none;border-bottom:2px solid transparent;color:#64748b;
         font-size:12px;font-weight:700;padding:10px 18px;cursor:pointer;transition:color .15s}
.tab-btn:hover{color:#94a3b8}
.tab-btn.active{color:#2dd4bf;border-bottom-color:#2dd4bf}
.sort-arrow{color:#2d3f5a;font-size:9px;margin-left:3px}
th:hover{background:#0d1525 !important;color:#e2e8f5 !important}
th:hover .sort-arrow{color:#2dd4bf}
.search-box{background:#0d1929; border:1px solid #1e2d45; color:#e2e8f5; padding:8px 14px;
            border-radius:6px; font-size:12px; outline:none; width:100%; min-width:280px; transition:border-color 0.2s;}
.search-box:focus{border-color:#2dd4bf;}
.search-box::placeholder{color:#475569;}
</style>
"""

UI_SCRIPT = r"""
<script>
// Filter table rows across Symbol, Company, and Sector (cells 0, 1, 2)
function filterTable() {
  var input = document.getElementById("searchBox").value.toLowerCase();
  var activeTab = document.querySelector('.tab.active');
  if (!activeTab) return;
  var rows = activeTab.querySelectorAll('tbody tr');
  rows.forEach(function(row) {
    if(row.cells.length >= 3) {
      var text = row.cells[0].textContent + " " + row.cells[1].textContent + " " + row.cells[2].textContent;
      row.style.display = text.toLowerCase().indexOf(input) > -1 ? "" : "none";
    }
  });
}

function switchTab(name, btn) {
  document.querySelectorAll('.tab').forEach(function(t){ t.classList.remove('active'); });
  document.querySelectorAll('.tab-btn').forEach(function(b){ b.classList.remove('active'); });
  var el = document.getElementById('tab-' + name);
  if (el) el.classList.add('active');
  btn.classList.add('active');
  filterTable(); // Re-apply global search to the new tab
}

var _ss = {};
function sortTable(th) {
  var tbl  = th.closest('table');
  var tb   = tbl.querySelector('tbody');
  var col  = parseInt(th.getAttribute('data-col'));
  var key  = tbl.id + '_' + col;
  var next = (_ss[key] === 'asc') ? 'desc' : 'asc';
  _ss[key] = next;
  tbl.querySelectorAll('th .sort-arrow').forEach(function(a){ a.textContent = '\u21c5'; });
  th.querySelector('.sort-arrow').textContent = next === 'asc' ? '\u25b2' : '\u25bc';
  var rows = Array.from(tb.querySelectorAll('tr'));
  var clean = function(v) { return v.replace(/[^0-9.+\-]/g,'').trim(); };
  rows.sort(function(a,b) {
    var ac = a.querySelectorAll('td')[col], bc = b.querySelectorAll('td')[col];
    if (!ac||!bc) return 0;
    var av = ac.textContent.trim(), bv = bc.textContent.trim();
    var an = parseFloat(clean(av)), bn = parseFloat(clean(bv));
    var cmp = (!isNaN(an)&&!isNaN(bn)) ? an-bn : av.localeCompare(bv,'en',{numeric:true});
    return next==='asc' ? cmp : -cmp;
  });
  // Strip zebra striping class handling to allow search filter logic to manage display
  rows.forEach(function(r,i){ r.style.background=i%2===0?'#0d1929':'#0f1c30'; tb.appendChild(r); });
}

document.addEventListener('click', function(e) {
  var btn = e.target.closest('.wl-btn');
  if (!btn) return;
  var syms  = JSON.parse(btn.getAttribute('data-syms') || '[]');
  var lines = [];
  for (var i=0; i<syms.length; i+=10) {
    lines.push(syms.slice(i,i+10).map(function(s){ return 'NSE:'+s; }).join(', '));
  }
  showWLModal(lines.join(',\n'), syms.length);
});

function showWLModal(content, count) {
  var old = document.getElementById('wl-modal'); if (old) old.remove();
  var m = document.createElement('div');
  m.id = 'wl-modal';
  m.style.cssText = 'position:fixed;inset:0;background:rgba(0,0,0,.85);z-index:9999;display:flex;align-items:center;justify-content:center;padding:20px';
  m.innerHTML = (
    '<div style="background:#0d1929;border:1px solid #2d3f5a;border-radius:10px;width:100%;max-width:700px;max-height:82vh;display:flex;flex-direction:column;box-shadow:0 24px 64px rgba(0,0,0,.9)">'
    +'<div style="display:flex;align-items:center;justify-content:space-between;padding:16px 20px;border-bottom:1px solid #1e2d45">'
    +'<div><div style="font-size:14px;font-weight:800;color:#2dd4bf">&#128203; TradingView Watchlist Export</div>'
    +'<div style="font-size:11px;color:#64748b;margin-top:2px">'+count+' symbols &middot; NSE: prefix added &middot; 10 per line</div></div>'
    +'<button onclick="document.getElementById(\'wl-modal\').remove()" style="background:none;border:none;color:#64748b;font-size:22px;cursor:pointer;padding:2px 8px">&times;</button></div>'
    +'<div style="padding:14px 20px;flex:1;overflow-y:auto">'
    +'<textarea id="wl-txt" readonly style="width:100%;height:250px;background:#060d18;border:1px solid #1e2d45;'
    +'color:#a5d6e8;font-family:Consolas,monospace;font-size:12px;line-height:1.9;padding:12px;border-radius:6px;resize:vertical;outline:none">'+content+'</textarea></div>'
    +'<div style="padding:14px 20px;border-top:1px solid #1e2d45;display:flex;gap:10px;align-items:center;flex-wrap:wrap">'
    +'<button onclick="cpWL()" style="background:#2dd4bf;color:#060d18;border:none;border-radius:5px;padding:9px 22px;font-size:12px;font-weight:800;cursor:pointer">&#8680; Copy to Clipboard</button>'
    +'<button onclick="dlWL()" style="background:rgba(45,212,191,.1);color:#2dd4bf;border:1px solid rgba(45,212,191,.3);border-radius:5px;padding:9px 16px;font-size:12px;font-weight:700;cursor:pointer">&#8595; Download .txt</button>'
    +'<span id="wl-msg" style="font-size:12px;color:#4ade80;font-weight:700;opacity:0;transition:opacity .3s"></span></div>'
    +'</div>'
  );
  document.body.appendChild(m);
  m.addEventListener('click', function(e){ if(e.target===m) m.remove(); });
}
function cpWL(){
  var ta=document.getElementById('wl-txt');
  ta.select(); ta.setSelectionRange(0,99999);
  try{ navigator.clipboard.writeText(ta.value); }catch(e){ document.execCommand('copy'); }
  var msg=document.getElementById('wl-msg');
  msg.textContent='\u2705 Copied!'; msg.style.opacity='1';
  setTimeout(function(){ msg.style.opacity='0'; },2500);
}
function dlWL(){
  var c=document.getElementById('wl-txt').value;
  var a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob([c],{type:'text/plain'}));
  a.download='vcp_watchlist.txt'; a.click();
}
</script>
"""

def grade_color(g):
    return {'A+':'#2dd4bf','A':'#4ade80','B':'#fbbf24','C':'#fb923c','REJECT':'#f87171'}.get(g,'#94a3b8')

def build_table_rows(stocks, bg1='#0d1929', bg2='#0f1c30'):
    rows = []
    for i, r in enumerate(stocks):
        bg = bg1 if i % 2 == 0 else bg2
        gc = grade_color(r['grade'])
        r30c = '#4ade80' if r['r30'] >= 30 else '#fbbf24' if r['r30'] >= 15 else '#f87171'
        r60c = '#4ade80' if r['r60'] >= 40 else '#fbbf24' if r['r60'] >= 20 else '#f87171'
        r90c = '#4ade80' if r['r90'] >= 50 else '#fbbf24' if r['r90'] >= 30 else '#f87171'
        dc   = '#4ade80' if r['drawdown'] >= -5 else '#fbbf24' if r['drawdown'] >= -15 else '#f87171'
        vc   = '#4ade80' if r['vol_contr'] <= -20 else '#fbbf24' if r['vol_contr'] < 0 else '#f87171'
        ac   = '#4ade80' if r['atr_ratio'] < 0.75 else '#fbbf24' if r['atr_ratio'] < 0.85 else '#f87171'
        tc   = '#4ade80' if r['tight_pct'] <= 8 else '#fbbf24' if r['tight_pct'] <= 12 else '#f87171'
        pp   = '<span style="color:#2dd4bf">✅</span>' if '✅' in r['pocket_pivot'] else '—'
        tv_link = f"{CFG['tv_chart_url']}?symbol=NSE:{r['symbol']}"

        rows.append(f"""<tr style="background:{bg};border-bottom:1px solid #1e2d45">
  <td style="padding:9px 12px;font-weight:800;font-size:12px;white-space:nowrap">
    <div style="display:flex;align-items:center;gap:8px">
        <a href="https://marketsmithindia.com/mstool/eval/{r['symbol'].lower()}/evaluation.jsp#/"
           target="_blank"
           style="color:#2dd4bf;text-decoration:none;border-bottom:1px dashed rgba(45,212,191,.4)"
           title="Open {r['symbol']} on MarketSmith India">{r['symbol']} ↗</a>
        <a href="{tv_link}" target="_blank"
           style="background:#1e2d45;color:#94a3b8;font-size:9px;padding:2px 5px;border-radius:4px;text-decoration:none;font-weight:bold;transition:all 0.2s"
           onmouseover="this.style.background='#2962ff';this.style.color='#fff'"
           onmouseout="this.style.background='#1e2d45';this.style.color='#94a3b8'"
           title="Open {r['symbol']} in TradingView">TV</a>
    </div>
  </td>
  <td style="padding:9px 8px;font-size:10px;color:#7cb4e0;max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{r['company']}</td>
  <td style="padding:9px 8px;font-size:10px;color:#64748b;white-space:nowrap">{r['sector']}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:#94a3b8">{r['market_cap']:,}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:12px;font-weight:700;color:#e2e8f5">₹{r['current']:,.2f}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:{r30c}">{r['r30']:+.1f}%</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:{r60c}">{r['r60']:+.1f}%</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:{r90c}">{r['r90']:+.1f}%</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:#60a5fa">{r['dist52']:+.1f}%</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:{dc}">{r['drawdown']:+.1f}%</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:{vc}">{r['vol_contr']:+.1f}%</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:{ac}">{r['atr_ratio']:.3f}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:{tc}">{r['tight_pct']:.1f}%</td>
  <td style="padding:9px 8px;font-size:11px;color:#a78bfa;white-space:nowrap">{r['pullbacks']}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:#fb923c">{r['rs_score']:.0f}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:#60a5fa">{r['vcp_score']:.0f}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:#a78bfa">{r['lead_score']:.0f}</td>
  <td style="padding:9px 8px;font-family:monospace;font-size:11px;color:#2dd4bf">{r['bos']}</td>
  <td style="padding:9px 12px;text-align:center">
    <span style="font-size:18px;font-weight:900;color:{gc}">{r['final_score']:.0f}</span>
  </td>
  <td style="padding:9px 8px;text-align:center">
    <span style="background:{gc}22;color:{gc};padding:2px 8px;border-radius:12px;
                 font-size:11px;font-weight:800;border:1px solid {gc}44">{r['grade']}</span>
  </td>
  <td style="padding:9px 8px;font-size:11px;white-space:nowrap">{r['setup']}</td>
  <td style="padding:9px 8px;text-align:center">{pp}</td>
</tr>""")
    return '\n'.join(rows)

TH = lambda t, i: (f'<th data-col="{i}" onclick="sortTable(this)" '
                   f'style="padding:8px 8px;text-align:left;font-size:9px;color:#64748b;'
                   f'letter-spacing:.7px;white-space:nowrap;background:#0a1220;'
                   f'cursor:pointer;user-select:none" '
                   f'title="Click to sort">{t} <span class=\'sort-arrow\'>⇅</span></th>')

HEADERS = ''.join(TH(h, i) for i, h in enumerate([
    'SYMBOL','COMPANY','SECTOR','MKTCAP(Cr)','PRICE',
    '30D%','60D%','90D%','DIST 52W','DRAWDOWN',
    'VOL CONTR','ATR RATIO','TIGHTNESS','VCP PULLS',
    'RS SCORE','VCP SCORE','LEAD SCORE','BREAKOUT RDY',
    'FINAL','GRADE','SETUP','PP'
]))

def generate_html(results, scan_time, total_scanned, errors, rejects):
    import json as _json
    elite   = [r for r in results if r['grade'] in ('A+','A')][:20]
    brk     = sorted([r for r in results if r['setup'] and 'Breakout' in r['setup']], key=lambda x: -x['final_score'])[:20]
    pp_list = [r for r in results if r['pocket_pivot'] and 'YES' in r['pocket_pivot']][:20]
    watch   = [r for r in results if r['grade'] == 'B'][:20]
    emerge  = [r for r in results if r['grade'] == 'C'][:20]

    def card(lbl, val, clr, sub):
        return (f'<div style="background:#0d1929;border:1px solid #1e2d45;border-radius:6px;padding:14px 16px">'
                f'<div style="font-size:9px;color:#64748b;letter-spacing:.8px;text-transform:uppercase;margin-bottom:6px">{lbl}</div>'
                f'<div style="font-size:26px;font-weight:900;color:{clr}">{val}</div>'
                f'<div style="font-size:10px;color:#64748b;margin-top:2px">{sub}</div></div>')

    cards = (card('Elite (A+/A)',    len(elite),   '#2dd4bf', 'Score &ge; 80') +
             card('Breakout Ready',  len(brk),     '#4ade80', 'Near pivot') +
             card('Pocket Pivots',   len(pp_list), '#fbbf24', "O'Neil confirmed") +
             card('Total Qualified', len(results), '#60a5fa', 'Passed all filters'))

    def tab_section(title, color, stocks, tid):
        if not stocks:
            return f'<div style="color:#475569;text-align:center;padding:48px;font-size:13px">No stocks in this category.</div>'
        rows = build_table_rows(stocks)
        syms = _json.dumps([r['symbol'] for r in stocks])
        safe_title = ''.join(c for c in title if ord(c) < 128)
        return (f'<div style="margin-bottom:36px">'
                f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;flex-wrap:wrap;gap:8px">'
                f'<h2 style="color:{color};font-size:13px;font-weight:800;letter-spacing:1px;text-transform:uppercase;'
                f'margin:0;padding:10px 16px;background:rgba(255,255,255,.03);border-left:3px solid {color};border-radius:0 4px 4px 0">'
                f'{title} &mdash; {len(stocks)} stocks</h2>'
                f'<button class="wl-btn" data-syms=\'{syms}\' data-title="{safe_title}" '
                f'style="background:rgba(45,212,191,.1);color:#2dd4bf;border:1px solid rgba(45,212,191,.3);'
                f'border-radius:5px;padding:7px 16px;font-size:11px;font-weight:700;cursor:pointer;white-space:nowrap">'
                f'&#128203; Export TradingView Watchlist</button></div>'
                f'<div style="overflow-x:auto;border:1px solid #1e2d45;border-radius:6px">'
                f'<table id="tbl-{tid}" style="width:100%;border-collapse:collapse;font-family:system-ui;min-width:1400px">'
                f'<thead><tr>{HEADERS}</tr></thead><tbody>{rows}</tbody></table></div></div>')

    s_elite  = tab_section('Elite VCP Candidates (A+ / A)', '#2dd4bf', elite,   'elite')
    s_brk    = tab_section('Breakout Ready',                '#4ade80', brk,     'brk')
    s_pp     = tab_section('Pocket Pivot Candidates',       '#fbbf24', pp_list, 'pp')
    s_watch  = tab_section('Watchlist (Grade B)',           '#a78bfa', watch,   'watch')
    s_emerge = tab_section('Emerging (Grade C)',            '#fb923c', emerge,  'emerge')
    s_all    = tab_section('All Qualified Stocks',          '#60a5fa', results, 'all')

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>VCP Scanner &mdash; {scan_time}</title>
{UI_CSS}
</head>
<body>
<div style="max-width:1700px;margin:0 auto;padding:24px">
  <div style="display:flex;align-items:flex-start;justify-content:space-between;
      margin-bottom:24px;padding-bottom:16px;border-bottom:1px solid #1e2d45;flex-wrap:wrap;gap:12px">
    <div>
      <h1 style="font-size:26px;font-weight:900;color:#2dd4bf;letter-spacing:-0.5px">&#128225; VCP SCANNER</h1>
      <p style="font-size:12px;color:#64748b;margin-top:4px">Minervini Trend Template &middot; Stage-2 &middot; VCP &middot; Pocket Pivot &middot; Breakout Ready</p>
    </div>
    <div style="text-align:right;font-size:11px;color:#64748b;line-height:1.9">
      <div>Run: <b style="color:#94a3b8">{scan_time}</b></div>
      <div>Qualified:<b style="color:#4ade80"> {len(results)}</b> &nbsp;|&nbsp; Rejected:<b style="color:#f87171"> {rejects}</b> &nbsp;|&nbsp; Errors:<b style="color:#fb923c"> {errors}</b></div>
      <div style="color:#334155">Yahoo Finance &middot; NSE &middot; Free &middot; No API Key</div>
    </div>
  </div>
  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:24px">{cards}</div>
  
  <div style="border-bottom:1px solid #1e2d45;margin-bottom:20px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;padding-bottom:8px">
    <div style="display:flex;gap:0;flex-wrap:wrap">
        <button class="tab-btn active" onclick="switchTab('elite',this)">&#127942; Elite (A+/A)</button>
        <button class="tab-btn" onclick="switchTab('brk',this)">&#128640; Breakout Ready</button>
        <button class="tab-btn" onclick="switchTab('pp',this)">&#9889; Pocket Pivots</button>
        <button class="tab-btn" onclick="switchTab('watch',this)">&#128064; Watchlist (B)</button>
        <button class="tab-btn" onclick="switchTab('emerge',this)">&#127807; Emerging (C)</button>
        <button class="tab-btn" onclick="switchTab('all',this)">&#128203; All Results</button>
    </div>
    <div>
        <input type="text" id="searchBox" class="search-box" placeholder="&#128269; Filter Symbol, Company, Sector..." onkeyup="filterTable()">
    </div>
  </div>

  <div id="tab-elite"  class="tab active">{s_elite}</div>
  <div id="tab-brk"    class="tab">{s_brk}</div>
  <div id="tab-pp"     class="tab">{s_pp}</div>
  <div id="tab-watch"  class="tab">{s_watch}</div>
  <div id="tab-emerge" class="tab">{s_emerge}</div>
  <div id="tab-all"    class="tab">{s_all}</div>
  
  <div style="margin-top:24px;padding-top:16px;border-top:1px solid #1e2d45;font-size:10px;color:#475569;text-align:center;line-height:2">
    VCP Score (60%) = Advance + Drawdown + Volume + ATR + Tightness &nbsp;|&nbsp; Leadership (40%) = 52W Proximity + Trend + RS Proxy + Liquidity<br>
    Yahoo Finance (15-min delayed) &middot; Run weekly after market close for Monday watchlist<br>
    <b>Not a buy recommendation. Always verify on TradingView + MarketSmith before acting.</b>
  </div>
</div>
{UI_SCRIPT}
</body>
</html>"""

def export_excel(results, outfile):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        cols = ['symbol','company','sector','market_cap','current',
                'r30','r60','r90','dist52','drawdown','vol_contr',
                'atr_ratio','tight_pct','pullbacks','rs_score',
                'vcp_score','lead_score','bos','final_score','grade','setup','pocket_pivot']
        df = pd.DataFrame([{c: r[c] for c in cols} for r in results])
        df.columns = ['Symbol','Company','Sector','MktCap(Cr)','Price',
                      '30D%','60D%','90D%','Dist52W%','Drawdown%','VolContr%',
                      'ATR Ratio','Tightness%','VCP Pulls','RS Score',
                      'VCP Score','Lead Score','Breakout Rdy','Final Score',
                      'Grade','Setup','Pocket Pivot']

        brk_syms = set(r['symbol'] for r in results if r['setup'] and 'Breakout' in r['setup'])
        pp_syms  = set(r['symbol'] for r in results if r['pocket_pivot'] and 'YES' in r['pocket_pivot'])

        tabs = [
            ('Elite (A+A)',     df[df['Grade'].isin(['A+','A'])]),
            ('Breakout Ready',  df[df['Symbol'].isin(brk_syms)]),
            ('Pocket Pivots',   df[df['Symbol'].isin(pp_syms)]),
            ('Watchlist (B)',   df[df['Grade'] == 'B']),
            ('Emerging (C)',    df[df['Grade'] == 'C']),
            ('All Results',     df),
        ]

        with pd.ExcelWriter(str(outfile), engine='openpyxl') as writer:
            for sheet_name, data in tabs:
                if not data.empty:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)

            wb = writer.book
            tab_colors = {
                'Elite (A+A)':    '2DD4BF',
                'Breakout Ready': '4ADE80',
                'Pocket Pivots':  'FBBF24',
                'Watchlist (B)':  'A78BFA',
                'Emerging (C)':   'FB923C',
                'All Results':    '60A5FA',
            }
            for ws in wb.worksheets:
                ws.freeze_panes = 'C2'
                clr = tab_colors.get(ws.title, '60A5FA')
                ws.sheet_properties.tabColor = clr
                
                ws.auto_filter.ref = ws.dimensions

                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 14
                ws.column_dimensions['A'].width = 16
                ws.column_dimensions['B'].width = 22
                ws.column_dimensions['C'].width = 18
                for cell in ws[1]:
                    cell.font      = Font(bold=True, color='FFFFFF', size=9)
                    cell.fill      = PatternFill('solid', fgColor='0D1929')
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
        return True
    except ImportError:
        print('  Warning: openpyxl not installed - skipping Excel (pip install openpyxl)')
        return False
    except Exception as e:
        print(f'  Warning: Excel export failed: {e}')
        return False

# ═══════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════
def main():
    interrupter = GracefulInterruptHandler()

    print('\n╔══════════════════════════════════════════════════════════════╗')
    print('║  VCP SCANNER  v2.4  —  Minervini · O\'Neil · Stage-2        ║')
    print('╚══════════════════════════════════════════════════════════════╝\n')
    print('💡 TIP: Press Ctrl+C at any time to instantly stop scanning and generate')
    print('the output HTML/Excel files for the stocks processed up to that point.\n')

    total     = len(NIFTY500)
    results   = []
    errors    = 0
    rejects   = 0
    scan_time = datetime.now().strftime('%d %b %Y  %I:%M %p')

    print(f'Scanning {total} NSE stocks via Yahoo Finance...\n')

    for i, sym in enumerate(NIFTY500, 1):
        if interrupter.interrupted:
            print(f'\nProceeding to generate report with the {len(results)} valid stocks found so far...\n')
            break

        try:
            df, info = fetch_data(sym)
            if df is None:
                print(f'  [{i:3d}/{total}] {sym:<15} — skip (no data)')
                continue

            result, reason = analyse(sym, df, info)

            if result is None:
                rejects += 1
                print(f'  [{i:3d}/{total}] {sym:<15} — ✗ {reason}')
                continue

            results.append(result)
            print(f'  [{i:3d}/{total}] {sym:<15} — vcp:{result["vcp_score"]:3.0f}'
                  f'  adv:{result["r90"]:+.0f}%'
                  f'  dd:{result["drawdown"]:+.0f}%'
                  f'  atr:{result["atr_ratio"]:.2f}'
                  f'  tight:{result["tight_pct"]:.1f}%'
                  f'  {result["setup"]}')

            time.sleep(CFG['delay_between_stocks'])

        except Exception as e:
            errors += 1
            print(f'  [{i:3d}/{total}] {sym:<15} — ERROR: {e}')
            continue

    if not results:
        print("\n❌ No stocks were fully processed to generate a report. Exiting.")
        sys.exit(0)

    print(f'\n═══ Computing RS Scores & Final Rankings ═══')

    results = compute_rs_scores(results)

    for r in results:
        r['rs_score']   = r.get('rs_score', 0)
        r['lead_score'] = leadership_score(r['prox_pts'], r['trend_pts'],
                                            r['rs_score'], r['avg_turn_liq'])
        r['final_score'] = composite_score(r['vcp_score'], r['lead_score'])
        r['grade']       = grade(r['final_score'])

    results.sort(key=lambda x: -x['final_score'])

    elite = [r for r in results if r['grade'] in ('A+','A')]
    watch = [r for r in results if r['grade'] == 'B']
    pps   = [r for r in results if '✅' in r['pocket_pivot']]

    print(f'\n╔══════════════════════════════════════════════════════════════╗')
    print(f'║  SCAN COMPLETE                                               ║')
    print(f'║  Qualified:{len(results):4d}  |  Elite:{len(elite):3d}  |  Rejected:{rejects:4d}  |  Errors:{errors:3d}  ║')
    print(f'╚══════════════════════════════════════════════════════════════╝\n')

    if elite:
        print('🏆 ELITE (A+/A):')
        for r in elite[:20]:
            print(f'   {r["symbol"]:<14} score:{r["final_score"]:5.1f}  '
                  f'grade:{r["grade"]}  vcp:{r["vcp_score"]:4.0f}  '
                  f'rs:{r["rs_score"]:4.0f}  {r["setup"]}')

    if pps:
        print(f'\n⚡ POCKET PIVOTS ({len(pps)}):')
        for r in pps[:10]:
            print(f'   {r["symbol"]:<14} score:{r["final_score"]:5.1f}')

    out_dir  = Path(__file__).resolve().parent
    html_out = out_dir / 'vcp_scanner_results.html'
    xlsx_out = out_dir / 'vcp_scanner_results.xlsx'

    print(f'\nGenerating HTML report...')
    html = generate_html(results, scan_time, total, errors, rejects)
    html_out.write_text(html, encoding='utf-8')
    print(f'✅ HTML saved: {html_out}')

    print(f'Generating Excel report...')
    if export_excel(results, xlsx_out):
        print(f'✅ Excel saved: {xlsx_out}')

    # print(f'\nOpening browser...\n')
    # webbrowser.open(html_out.as_uri())


if __name__ == '__main__':
    main()